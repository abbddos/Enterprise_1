import psycopg2

#from werkzeug.utils import secure_filename
import os
from openpyxl import load_workbook
import random
from datetime import date
import pandas as pd
import EnterpriseConfig
import hashlib


ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'xls','xlsx'])
# Creating connection engine using Root connection to handle administrative
# functions such as loging in, sending emails in case of missing passords, and
# changing users' passwords.

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def root():
    conn = psycopg2.connect("dbname = enterprise user = {} password = {} port = 5432".format(EnterpriseConfig.dbusername, EnterpriseConfig.dbpassword))
    cur = conn.cursor()
    return conn, cur

def connector(usrname, passwd):
    conn = psycopg2.connect("dbname = enterprise user = {} password = {} port = 5432".format(usrname, passwd))
    cur = conn.cursor()
    return conn, cur

def NewUser(newname):
    con, cur = root()
    i = 1
    New_User = newname
    cur.execute('SELECT username FROM users')
    AllUsers = cur.fetchall()
    for oneuser in AllUsers:
        while New_User == oneuser[0]:
            New_User = newname + str(i)
            i += 1
    con.close()
    return New_User

# Logger funtions uses root connection Engine to search for username and password
# If found Logger value returns the boolean value of True, otherwise, it returns
# False.
def Logger(usrname, passwd):
    passwd1 = passwd
    if passwd != 'admin':
        m = hashlib.sha256()
        m.update(passwd.encode('utf8'))
        passwd1 = m.hexdigest()
        con, cur = root()
        cur.execute('SELECT username, password, usertype, status FROM users WHERE username = %s and password = %s', (usrname, passwd1))
        result = cur.fetchone()
        try:
            if result[0] == usrname and result[1] == passwd1 and result[3] == 'Active':
                return {'username': result[0], 'password': result[1], 'logged': True, 'role':result[2]}
            else:
                return {'logged': False}
        except:
            return {'logged': False}
    else:
        con, cur = root()
        cur.execute('SELECT username, password, usertype FROM users WHERE username = %s and password = %s', (usrname, passwd1))
        result = cur.fetchone()
        try:
            if result[0] == usrname and result[1] == passwd1:
                return {'username': result[0], 'password': result[1], 'logged': True, 'role': result[2]}
        except:
            return {'logged': False}

# .......The following functions are specifically for the user to update his/her own...
# .......information and password.

def GetUser(user):
    con, cur = root()
    cur.execute('SELECT id, firstname, lastname, position, department, email, phone1, phone2, usertype FROM users WHERE username = %s', (user,))
    usrdata = cur.fetchall()
    con.close()
    return usrdata


def UpdateProfile(firstname, lastname, email, phone1, phone2, user):
    con, cur = root()
    cur.execute('UPDATE users SET firstname = %s, lastname = %s, email = %s, phone1 = %s, phone2 = %s WHERE username = %s',(firstname, lastname, email, phone1, phone2, user))
    con.commit()
    con.close()

def ChangePassword(user, currentpswd, newpswd):
    m = hashlib.sha256()
    n = hashlib.sha256()
    newpass = None
    oldpass = currentpswd
    if currentpswd != 'admin':
        m.update(newpswd.encode('utf8'))
        newpass = m.hexdigest()
        n.update(currentpswd.encode('utf8'))
        oldpass = n.hexdigest()
    else:
        m.update(newpswd.encode('utf8'))
        newpass = m.hexdigest()
    con, cur = root()
    cur.execute('UPDATE users SET password = %s WHERE username = %s AND password = %s', (newpass, user, oldpass))
    cur.execute("ALTER ROLE {} WITH LOGIN PASSWORD '{}'".format(user, newpass))
    con.commit()
    con.close()

def ChangePassword1(usr, newpswd):
     m = hashlib.sha256()
     m.update(newpswd.encode('utf8'))
     newpass = m.hexdigest()

     con, cur = root()
     cur.execute('UPDATE users SET password = %s WHERE username = %s', (newpass, usr))
     cur.execute("ALTER ROLE {} WITH LOGIN PASSWORD '{}'".format(usr, newpass))
     con.commit()
     con.close()
# .......The following functions are for Administrators to create users, get users info,
# .......update users, and activate and deactivate users..
# .......the GetUsers function gets information from all users from database in order to...
# .......display that information in organized tables, and it's different from GetUser function above..

def GetUsers():
    con, cur = root()
    cur.execute('SELECT id, firstname, lastname, username, company, position, department, email, phone1, phone2, usertype, status FROM users ORDER BY id')
    usrdata = cur.fetchall()
    con.close()
    return usrdata

def GetApprovals(username):
    con, cur = root()
    cur.execute('SELECT can_approve FROM approvers WHERE username = %s', (username,))
    data = cur.fetchall()
    return data

def GetUserInfo(id):
    con, cur = root()
    cur.execute('SELECT * FROM users WHERE id = %s', (id,))
    data = cur.fetchall()
    con.close()
    return data

def UpdateUser(sess_uname, sess_pswd, id, firstname, lastname, company, position, department, email, phone1, phone2, usertype, status, approvals, username):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE users SET firstname = %s, lastname = %s, company = %s, position = %s, department = %s, email = %s, phone1 = %s, phone2 = %s, usertype = %s, status = %s WHERE id = %s',
    (firstname, lastname, company, position, department, email, phone1, phone2, usertype, status, id))
    cur.execute('DELETE FROM approvers WHERE username = %s', (username,))
    con.commit()
    for a in range(len(approvals)):
        cur.execute('INSERT INTO approvers(firstname, lastname, username, position, department, email, can_approve) VALUES(%s, %s, %s, %s, %s, %s, %s)', (firstname, lastname, username, position, department, email, approvals[a]))
        con.commit()
    con.close()

def CreateUser(sess_uname, sess_pswd, firstname, lastname, company, position, department, email, phone1, phone2, usrtype, approvals):
    con, cur = connector(sess_uname, sess_pswd)
    #... Adding user to users table...
    UserName = firstname[0].lower() + lastname.lower()
    NewName = NewUser(UserName)
    passwd = NewName + '@123'
    m = hashlib.sha256()
    m.update(passwd.encode('utf8'))
    Password = m.hexdigest()
    cur.execute('INSERT INTO users(firstname, lastname, username, password, company, position, department, email, phone1, phone2, usertype) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
    (firstname, lastname, NewName, Password, company, position, department, email, phone1, phone2, usrtype))
    #... Creating Database User with UserType permissions.
    cur.execute("CREATE ROLE {} WITH SUPERUSER LOGIN PASSWORD '{}'".format(NewName, Password))
    con.commit()
    for a in range(len(approvals)):
        cur.execute('INSERT INTO approvers(firstname, lastname, username, position, department, email, can_approve) VALUES(%s, %s, %s, %s, %s, %s, %s)', (firstname, lastname, NewName, position, department, email, approvals[a]))
        con.commit()
    con.close()
    return NewName, passwd

def CreateMultipleUsers(sess_uname, sess_pswd, FileName):
    wb = load_workbook(filename = FileName)
    ws = wb['Sheet1']
    i = 2
    while ws.cell(row = i, column = 1).value is not None:
        firstname = str(ws.cell(row = i, column = 1).value)
        lastname = str(ws.cell(row = i, column = 2).value)
        company = str(ws.cell(row = i, column = 3).value)
        position = str(ws.cell(row = i, column = 4).value)
        department = str(ws.cell(row = i, column = 5).value)
        email = str(ws.cell(row = i, column = 6).value)
        phone1 = str(ws.cell(row = i, column = 7).value)
        phone2 = str(ws.cell(row = i, column = 8).value)
        type = str(ws.cell(row = i, column = 9).value)
        approvals = []
        CreateUser(sess_uname, sess_pswd, firstname, lastname, company, position, department, email, phone1, phone2, type, approvals)
        i += 1

def ResetPassword(sess_uname, sess_pswd, user):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT firstname, username, email FROM users WHERE id = %s', (user,))
    name = cur.fetchone()
    rst_pass = str(name[1]) + '@123'
    m = hashlib.sha256()
    m.update(rst_pass.encode('utf8'))
    resetpass = m.hexdigest()
    cur.execute('UPDATE users SET password = %s WHERE id = %s' ,(resetpass, user))
    cur.execute("ALTER ROLE {} WITH LOGIN PASSWORD '{}'".format(name[1], resetpass))
    con.commit()
    con.close()
    return rst_pass, name 


def GetCompanyProfile():
    con, cur = root()
    cur.execute("SELECT * FROM companyprofile WHERE id = 1")
    data = cur.fetchone()
    con.close()
    return data

def UpdateCompanyProfile(sess_uname, sess_pswd, name, address, phone1, phone2, email, pobox, registration, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE companyprofile SET companyname = %s, address = %s, phone_1 = %s, phone_2 = %s, email = %s, pobox = %s, registration = %s, description = %s',
    (name, address, phone1, phone2, email, pobox, registration, description))
    con.commit()
    con.close()

def GetProviders():
    con, cur = root()
    cur.execute('SELECT id, name, address, phone_1, phone_2, email, pobox FROM providers')
    data = cur.fetchall()
    con.close()
    return data

def CreateProvider(sess_uname, sess_pswd, name, address, phone1, phone2, email, pobox, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT CreateProvider(%s, %s, %s, %s, %s, %s, %s)',
    (name, address, phone1, phone2, email, pobox, description))
    con.commit()
    con.close()

def FetchProvider(sess_uname, sess_pswd, prv):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT * FROM providers WHERE id = %s', (prv,))
    data = cur.fetchall()
    con.close()
    return data

def UpdateProvider(sess_uname, sess_pswd, prv, name, address, phone1, phone2, email, pobox, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE providers SET name = %s, address = %s, phone_1 = %s, phone_2 = %s, email = %s, pobox = %s, description = %s WHERE id = %s',
    (name, address, phone1, phone2, email, pobox, description, prv))
    con.commit()
    con.close()

def ProvidersList():
    con, cur = root()
    cur.execute('SELECT Name FROM Providers')
    provs = cur.fetchall()
    choices = []
    for i in provs:
        choices.append((i[0],i[0]))
    con.close()
    return choices

def GroupList():
    con, cur = root()
    cur.execute('SELECT id, groupname FROM grp ORDER BY id')
    data = cur.fetchall()
    con.close()
    return data

def Groups():
    con, cur = root()
    cur.execute('SELECT groupname FROM grp')
    data = cur.fetchall()
    choices = []
    for i in data:
        choices.append((i[0],i[0]))
    con.close()
    return choices

def GetGroup(id):
    con, cur = root()
    cur.execute('SELECT * FROM grp WHERE id = %s', (id,))
    data = cur.fetchall()
    con.close()
    return data

def AddGroup(sess_uname, sess_pswd, name, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO grp(groupname, description) VALUES(%s, %s)', (name, description))
    con.commit()
    con.close()

def UpdateGroup(sess_uname, sess_pswd, id, name, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE grp SET groupname = %s, description = %s WHERE id = %s', (name, description, id))
    con.commit()
    con.close()

def CreateItem(sess_uname, sess_pswd, itemname, brand, provider, unit, uprice, ucost, description, size, color, sku, partnum, ieme, length, width, height, diameter, lunit, wunit, hunit, dunit, grp, category, secondaryunit):
    code = random.randint(100000000000,999999999999)
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT CreateItem(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
    (str(code), itemname, brand, provider, unit, uprice, ucost, description, size, color, sku, partnum, ieme, length, width, height, diameter, lunit, wunit, hunit, dunit, grp, category, secondaryunit))
    con.commit()
    con.close()

def UpdateItem(sess_uname, sess_pswd, itm, itemname, brand, provider, unit, uprice, ucost, description, size, color, sku, partnum, ieme, length, width, height, diameter, lunit, wunit, hunit, dunit, grp, category, secondaryunit):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT UpdateItem(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
    (itm, itemname, brand, provider, unit, uprice, ucost, description, size, color, sku, partnum, ieme, length, width, height, diameter, lunit, wunit, hunit, dunit, grp, category, secondaryunit))
    con.commit()
    con.close()

def GetItems():
    con, cur = root()
    cur.execute('SELECT * FROM Items ORDER BY ID')
    data = cur.fetchall()
    con.close()
    return data

def FetchItem(itm):
    con, cur = root()
    cur.execute('SELECT * FROM items WHERE id = %s', (itm,))
    data = cur.fetchall()
    con.close()
    return data

def ItemPicker():
    con, cur = root()
    cur.execute('SELECT code, item, provider, unit FROM items')
    data = cur.fetchall()
    con.close()
    return data

def ItemAdder(code):
    con, cur = root()
    cur.execute('SELECT code, item, unit, unit_price, unit_cost FROM items WHERE code = %s', (code,))
    data = cur.fetchall()
    con.close()
    return data

def GetPackages():
    con, cur = root()
    cur.execute('SELECT packagecode, packagename FROM packages GROUP BY packagecode, packagename ORDER BY packagecode')
    data = cur.fetchall()
    con.close()
    return data

def CreatePackage(sess_uname, sess_pswd, packagename, itemcode, itemname, unit, unitprice, unitcost, quantity, description):
    code = 'pkg_' + str(random.randint(100000000000,999999999999))
    con, cur = connector(sess_uname, sess_pswd)
    for i in range(len(itemcode)):
        cur.execute('INSERT INTO packages(packagecode, packagename, itemcode, itemname, unit, unit_price, unit_cost, quantity, description) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s)',
        (code, packagename, itemcode[i], itemname[i], unit[i], unitprice[i], unitcost[i], quantity[i], description))
        con.commit()

    con.close()

def FetchPackage(pkg):
    con, cur =root()
    #...Get Package Metadata, Pakage code, Name, ad Description.
    cur.execute('SELECT packagecode, packagename, description FROM packages WHERE packagecode = %s GROUP BY packagecode, packagename, description', (pkg,))
    data1 = cur.fetchone()
    #...Get Package content.
    cur.execute('SELECT itemcode, itemname, unit, unit_price, unit_cost, quantity FROM packages WHERE packagecode = %s', (pkg,))
    data2 = cur.fetchall()
    con.close()
    return data1, data2

def UpdatePackage(sess_uname, sess_pswd, pkg, packagename, itemcode, itemname, unit, unitprice, unitcost, quantity, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('DELETE FROM packages WHERE packagecode = %s', (pkg,))
    for i in range(len(itemcode)):
        cur.execute('INSERT INTO packages(packagecode, packagename, itemcode, itemname, unit, unit_price, unit_cost, quantity, description) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s)',
        (pkg, packagename, itemcode[i], itemname[i], unit[i], unitprice[i], unitcost[i], quantity[i], description))
        con.commit()
    con.close()

def PackagePicker():
    con, cur = root()
    cur.execute('SELECT packagecode, packagename FROM packages GROUP BY packagecode, packagename ORDER BY packagecode')
    data = cur.fetchall()
    con.close()
    return data

def PackageAdder(code):
    con, cur = root()
    cur.execute('SELECT packagecode, packagename FROM packages WHERE packagecode = %s GROUP BY packagecode, packagename ORDER BY packagecode', (code,))
    data = cur.fetchall()
    con.close()
    return data

def CreateSecondaryUnit(sess_uname, sess_pswd,name, code, unit, measure):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO SecondaryUnits(Name, Code, Unit, Measure) VALUES(%s, %s, %s, %s)', (name, code, unit, measure))
    con.commit()
    con.close()

def GetSecondaryUnits():
    con, cur = root()
    cur.execute('SELECT * FROM SecondaryUnits ORDER BY code')
    data = cur.fetchall()
    con.close()
    return data

def GrabSecondaryUnit(code):
    con, cur = root()
    cur.execute('SELECT * FROM SecondaryUnits WHERE code = %s', (code,))
    data = cur.fetchone()
    con.close()
    return data

def SecondaryUnits():
    con, cur = root()
    cur.execute('SELECT code FROM SecondaryUnits')
    data = cur.fetchall()
    choices = []
    for i in data:
        choices.append((i[0],i[0]))
    con.close()
    return choices

def UpdateSecondaryUnit(sess_uname, sess_pswd, name, code, unit, measure):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE SecondaryUnits SET name = %s, unit = %s, measure = %s WHERE code = %s',(name, unit, measure, code))
    con.commit()
    con.close()

def FetchSecondaryUnits():
    con, cur = root()
    cur.execute('SELECT code FROM SecondaryUnits')
    data = cur.fetchall()
    con.close()
    return data

def GetWareHouses():
    con, cur = root()
    cur.execute('SELECT code, name FROM warehouses ORDER BY code')
    data = cur.fetchall()
    con.close()
    return data

def FetchWarehouse(code):
    con, cur = root()
    cur.execute('SELECT location, code, name, description FROM warehouses WHERE code = %s', (code,))
    data = cur.fetchall()
    con.close()
    return data

def CreateWarehouse(sess_uname, sess_pswd, Name, Code, Location, Description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO warehouses(location, code, name, description) VALUES(%s, %s, %s, %s)', (Location, Code, Name, Description))
    con.commit()
    con.close()

def UpdateWarehouse(sess_uname, sess_pswd, code, Name, Location, Description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE warehouses SET location = %s, name = %s, description = %s WHERE code = %s', (Location, Name, Description, code))
    con.commit()
    con.close()

def CreateBin(sess_uname, sess_pswd, name, code, wh, status, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO bins(code, name, warehouse, description, status) VALUES(%s, %s, %s, %s, %s)', (code, name, wh, description, status))
    con.commit()
    con.close()

def GetBins(code):
    con, cur = root()
    cur.execute('SELECT code, name, status FROM bins WHERE warehouse = %s', (code,))
    data = cur.fetchall()
    con.close()
    return data

def GrabBins(code):
    con, cur = root()
    cur.execute('SELECT code FROM bins WHERE warehouse = %s', (code,))
    data = cur.fetchall()
    con.close()
    return data

def BinInfo(code):
    con, cur = root()
    cur.execute('SELECT code, name, description, status FROM bins WHERE code = %s', (code,))
    data = cur.fetchone()
    con.close()
    return data

def UpdateBin(sess_uname, sess_pswd, code, name, status, desc):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE bins SET name = %s, description = %s, status = %s WHERE code = %s', (name, desc, status, code))
    con.commit()
    con.close()

def InboundTransaction(sess_uname, sess_pswd, ItemName, ItemCode, Warehouse, Bin, Unit, Qty, Status, Comments):
    con, cur = connector(sess_uname, sess_pswd)
    code = 'Inb_' + str(date.today()) + '-' + str(random.randint(100000,999999))
    for i in range(len(ItemCode)):
        cur.execute('SELECT Inbound(%s, %s, %s, %s, %s, %s, %s, %s, %s)',(code, sess_uname, ItemName[i], ItemCode[i], Warehouse, Bin[i], Unit[i], Qty[i], Status))
        cur.execute('UPDATE transactions SET comments  = %s WHERE transid = %s',(Comments, code))
        con.commit()
    con.close()
    return code

def OutboundTransaction(sess_uname, sess_pswd, ItemName, ItemCode, Warehouse, Bin, Unit, Qty, Status, Comments):
    con, cur = connector(sess_uname, sess_pswd)
    code = 'Out_' + str(date.today()) + '-' +str(random.randint(100000,999999))
    for i in range(len(ItemCode)):
        cur.execute('SELECT Outbound(%s, %s, %s, %s, %s, %s, %s, %s, %s)',(code, sess_uname, ItemName[i], ItemCode[i], Warehouse, Bin[i], Unit[i], Qty[i], Status))
        cur.execute('UPDATE transactions SET comments = %s WHERE transid = %s',(Comments, code))
        con.commit()
    con.close()
    return code

def GetTrans():
    con, cur = root()
    cur.execute('SELECT createdon, createdby, editedon, editedby, type, warehouse, status, transid FROM transactions GROUP BY transid, createdon, createdby, editedon, editedby, type, warehouse, status')
    data = cur.fetchall()    
    con.close()
    return data

def TransInvoice(transid):
    con, cur = root()
    cur.execute('SELECT itemcode, itemname, bin, unit, quantity FROM transactions where transid = %s', (transid,))
    data1 = cur.fetchall()
    cur.execute('SELECT createdon, editedon, comments FROM transactions WHERE transid = %s GROUP BY createdon, editedon, comments', (transid,))
    data2 = cur.fetchone()
    con.close()
    return data1, data2

def TransInfo(transid):
    con, cur = root()
    cur.execute('SELECT type, warehouse, status, transid, comments, reference FROM transactions WHERE transid = %s GROUP BY transid, type, warehouse, status, comments, reference',(transid,))
    data1 = cur.fetchone()
    cur.execute('SELECT itemcode, itemname, bin, unit, quantity FROM transactions WHERE transid = %s', (transid,))
    data2 = cur.fetchall()
    con.close()
    return data1, data2

def UpdateTransaction(sess_uname, sess_pswd, transid, newstatus, comments):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT updatetransaction(%s, %s, %s)', (transid, sess_uname, newstatus))
    cur.execute('UPDATE transactions SET comments = %s WHERE transid = %s', (comments, transid))
    con.commit()
    con.close()

def GetRequests():
    con, cur = root()
    cur.execute('SELECT rqstid, createdon, createdby, editedon, editedby, type, status FROM request GROUP BY rqstid, createdon, createdby, editedon, editedby, type, status')
    data = cur.fetchall()
    con.close()
    return data

def InvApprovers():
    con, cur = root()
    cur.execute('SELECT firstname, lastname, username FROM approvers group by firstname, lastname, username')
    data = cur.fetchall()
    con.close()
    return data
    
def CreateRequest(sess_uname, sess_pswd, type, status, approver, comments, code, name, unit, quantity):
    con, cur = connector(sess_uname, sess_pswd)
    rqstcode = ''

    if type == 'Inbound':
        rqstcode = 'req_inb_' + str(date.today()) + '_' + str(random.randint(100000, 999999))
    elif type == 'Outbound':
        rqstcode = 'req_out_' + str(date.today()) + '_' + str(random.randint(100000, 999999))
    
    for i in range(len(code)):
        cur.execute('INSERT INTO request(rqstid, createdon, createdby, type, itemcode, itemname, unit, quantity, status, comments, approver) VALUES(%s, now(), %s, %s, %s, %s, %s, %s, %s, %s, %s)', 
        (rqstcode, sess_uname, type, code[i], name[i], unit[i], quantity[i], status, comments, approver))
        con.commit()
    cur.execute('SELECT firstname, username, email FROM approvers WHERE username = %s GROUP BY firstname, username, email', (approver,))
    data = cur.fetchone()
    con.close()
    return data, rqstcode

def UpdateRequest(sess_uname, sess_pswd, status, comment, reqid):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE request SET editedby = %s, editedon = now(), status = %s, comments = %s WHERE rqstid = %s', (sess_uname, status, comment, reqid))
    con.commit()
    con.close()

def ReqInfo(reqid):
    con, cur = root()
    cur.execute('SELECT rqstid, type, status, comments, approver FROM request WHERE rqstid = %s GROUP BY rqstid, type, status, comments, approver', (reqid,))
    data1 = cur.fetchone()
    cur.execute('SELECT itemcode, itemname, unit, quantity FROM request WHERE rqstid = %s', (reqid,))
    data2 = cur.fetchall()
    cur.execute('SELECT firstname, lastname, username FROM approvers WHERE username = %s GROUP BY firstname, lastname, username', (data1[4],))
    data3 = cur.fetchone()
    con.close()
    return data1, data2, data3

def ApprRequests():
    List = []
    con, cur = root()
    cur.execute("SELECT rqstid FROM request WHERE status = 'Approved' GROUP BY rqstid")
    data1 = cur.fetchall()
    cur.execute("SELECT invoicecode FROM invoices WHERE invstatus = 'Approved' GROUP BY invoicecode")
    data2 = cur.fetchall()
    con.close()
    for d1 in data1:
        List.append(d1)
    for d2 in data2:
        List.append(d2)
    return List

def Warehouse_Inventory_to_CSV(sess_uname, sess_pswd, fromdate, todate, whs):
    con, cur = connector(sess_uname, sess_pswd)
    data = pd.read_sql('SELECT itemcode, itemname, warehouse, bin, unit, quantity, lastupdate, unitprice, bulkprice FROM inventory WHERE warehouse = ANY(ARRAY{}::VARCHAR[]) ORDER BY warehouse, lastupdate'.format(whs), con)
    file = data.to_csv()
    con.close()
    return file

def Warehouse_Inventory_to_PDF(sess_uname, sess_pswd, fromdate, todate, whs):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT itemcode, itemname, warehouse, bin, unit, quantity, lastupdate, unitprice, bulkprice FROM inventory WHERE warehouse = ANY(ARRAY{}::VARCHAR[]) ORDER BY warehouse, lastupdate'.format(whs))
    data = cur.fetchall()
    con.close()
    return data

def Items_Packs_Inventory_to_CSV(sess_uname, sess_pswd, fromdate, todate, itm, pkg):
    con, cur = connector(sess_uname, sess_pswd)
    data = pd.read_sql('SELECT itemcode, itemname, warehouse, bin, unit, quantity, lastupdate, unitprice, bulkprice FROM inventory WHERE itemcode = ANY(ARRAY{}::VARCHAR[]) OR itemcode = ANY(ARRAY{}::VARCHAR[]) ORDER BY itemcode, lastupdate'.format(itm, pkg), con)
    file = data.to_csv()
    con.close()
    return file

def Items_Packs_Inventory_to_PDF(sess_uname, sess_pswd, fromdate, todate, itm, pkg):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT itemcode, itemname, warehouse, bin, unit, quantity, lastupdate, unitprice, bulkprice FROM inventory WHERE itemcode = ANY(ARRAY{}::VARCHAR[]) OR itemcode = ANY(ARRAY{}::VARCHAR[]) ORDER BY itemcode, lastupdate'.format(itm, pkg))
    data = cur.fetchall()    
    con.close()
    return data

def Transactions_Report_to_CSV(sess_uname, sess_pswd):
    con, cur = connector(sess_uname, sess_pswd)
    data = pd.read_sql('SELECT * FROM TRANSACTIONS ORDER BY id', con)
    file = data.to_csv()
    con.close()
    return file

def Transactions_Report_to_PDF(sess_uname, sess_pswd):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT * FROM transactions ORDER BY status')
    data1 = cur.fetchall()
    con.close()
    return data1

def Bins_Report_to_CSV(sess_uname, sess_pswd, whs):
    con, cur = connector(sess_uname, sess_pswd)
    data = pd.read_sql('SELECT * FROM Bins_View WHERE warehouse = ANY(ARRAY{}::VARCHAR[])'.format(whs), con)
    file = data.to_csv()
    con.close()
    return file

def Bins_Report_to_PDF(sess_uname, sess_pswd, whs):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT * FROM Bins_View WHERE warehouse = ANY(ARRAY{}::VARCHAR[])'.format(whs))
    data = cur.fetchall()
    con.close()
    return data

#ApprRequests()
